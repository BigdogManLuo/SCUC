import numpy as np
from io import StringIO
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def read_txt(file_path,is_storage_price=False):

    with open(file_path, 'r', encoding='ANSI') as file:
        lines = file.readlines()

    #去掉开头的"//"和空格的影响
    lines = [line.strip().lstrip('//') for line in lines]
    
    #修复储能报价和容量的表头格式问题
    if is_storage_price and ('机组 报价段'in lines[0]) or ('机组\t报价段'in lines[0]):
        lines[0] = lines[0].replace('机组 报价段', '机组\t报价段-1\t报价段0\t报价段1\t报价段2\t报价段3\t报价段4')

    elif is_storage_price and '价格'in lines[0]:
        lines[0] ='机组序号\t价格-1\t价格0\t价格1\t价格2\t价格3\t价格4\t价格5'

    #修复unitdata.txt文件中的表头格式问题
    if 'unitdata.txt' in file_path:
        if '下爬坡率 (MW/h)' in lines[0]:
            lines[0] = lines[0].replace('下爬坡率 (MW/h)', '下爬坡率(MW/h)')
        elif '下爬坡率\t(MW/h)' in lines[0]:
            lines[0] = lines[0].replace('下爬坡率\t(MW/h)', '下爬坡率(MW/h)')

    return lines

def txt_to_dataframe(lines):

    cleaned_lines = []
    for line in lines:
        cleaned_line = '\t'.join(line.split())
        cleaned_lines.append(cleaned_line)

    cleaned_data = "\n".join(cleaned_lines)

    df = pd.read_csv(StringIO(cleaned_data), delimiter='\t')

    return df

def readStorageBidInfo(file_path):

    lines=read_txt(file_path)
    
    cleaned_lines = lines[1:]
    cleaned_lines = "\t".join(cleaned_lines)
    data = list(map(int, cleaned_lines.split()))
    
    data = data[1:]

    return data

def parse_log_file(file_path):

    with open(file_path, 'r', encoding='ANSI') as file:
        content = file.read()

    # 使用正则表达式提取每个数据块
    blocks = {
        'BranchUnitSensi': re.search(r'<BranchUnitSensi::dky type=全数>(.*?)</BranchUnitSensi::dky>', content, re.DOTALL),
        'BranchData': re.search(r'<BranchData::dky type=全数>(.*?)</BranchData::dky>', content, re.DOTALL),
        'BranchBase': re.search(r'<BranchBase::dky type=全数>(.*?)</BranchBase::dky>', content, re.DOTALL)
    }

    df_tuple = ()

    # 提取每个数据块并转换为DataFrame
    for block_name, block in blocks.items():
        if block:
            block_content = block.group(1).strip()
            lines = block_content.splitlines()
            df=txt_to_dataframe(lines)
            df = df.loc[:, ~df.columns.str.contains('//')]
            df_tuple += (df,)

    return df_tuple

def getSegmentedCostInfo(prices,deltaP,P_min):
   
    # 计算每个区间的斜率和截距
    a = []
    b = []
    for i in range(len(prices) - 1):
        # 计算每个区间的斜率
        slope = (prices[i+1] - prices[i]) / deltaP  # 斜率
        intercept = prices[i] - slope * (P_min + i * deltaP)  # 截距
        a.append(slope)
        b.append(intercept)

    return a,b


def getSegmentedPoints(num_seg,deltaP,P_min):

    # 计算每段区间的左右端点

    Pmins=[]
    Pmaxs=[]
    
    for i in range(num_seg):
        Pmins.append(P_min+i*deltaP)
        Pmaxs.append(P_min+(i+1)*deltaP)

    return Pmins,Pmaxs


def addColors(filePath,bin_table_name=['机组状态','机组启动状态','机组停机状态','储能状态'],continuous_table_name=['机组功率']):
    
    
    # 加载工作簿
    wb = load_workbook(filePath)

    # 定义填充样式
    fill_1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
    fill_0 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色


    for sheet_name in bin_table_name:

        ws = wb[sheet_name]

        # 更改单元格底色
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                if cell.value == 1:
                    cell.fill = fill_1
                elif cell.value == 0:
                    cell.fill = fill_0


    # 为连续值的变量设置颜色渐变
    for sheet_name in continuous_table_name:

        ws = wb[sheet_name]

        color_scale_rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                        mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                        end_type='max', end_color='FF0000')
        ws.conditional_formatting.add(f'B2:Z{ws.max_row}', color_scale_rule)


    # 保存工作簿
    wb.save(filePath)
    
    
def Sols2Excel(U_unit,P_unit,U_ch,U_dch,U_ES,P_ES_ch,P_ES_dch,instance_num,is_opt):
    
    #检测路径，如果不存在则创建
    if not os.path.exists(f'results/{instance_num}'):
        os.makedirs(f'results/{instance_num}')

    eps=1e-2
    
    N_ESs=U_ch.shape[0]
    N_units=U_unit.shape[0]

    U_unit=pd.DataFrame(U_unit)
    U_unit['机组编号']=U_unit.index+1
    cols = list(U_unit)
    cols.insert(0, cols.pop(cols.index('机组编号')))
    U_unit = U_unit.loc[:, cols]

    P_unit=pd.DataFrame(P_unit)
    P_unit['机组编号']=P_unit.index+1
    cols = list(P_unit)
    cols.insert(0, cols.pop(cols.index('机组编号')))
    P_unit = P_unit.loc[:, cols]

    storage_status=5*np.ones((N_ESs,24))
    storage_power=5*np.ones((N_ESs,24))
    for i in range(N_ESs):
        for t in range(24):
            if abs(U_ch[i,t])-1<eps:
                storage_status[i,t]=-1
                storage_power[i,t]=-P_ES_ch[i,t]
            elif abs(U_ch[i,t])-1<eps:
                storage_status[i,t]=1
                storage_power[i,t]=P_ES_dch[i,t]
            elif abs(U_ES[i,t]-1)<eps:
                storage_status[i,t]=0
                storage_power[i,t]=0
            else:
                raise ValueError("Invalid storage status")
    
    ES_status = pd.DataFrame(storage_status)
    ES_status['储能编号'] = ES_status.index+1
    cols = list(ES_status)
    cols.insert(0, cols.pop(cols.index('储能编号')))
    ES_status = ES_status.loc[:, cols]

    ES_power = pd.DataFrame(storage_power)
    ES_power['储能编号'] = ES_power.index+1
    cols = list(ES_power)
    cols.insert(0, cols.pop(cols.index('储能编号')))
    ES_power = ES_power.loc[:, cols]
    
    if is_opt:
        filepath=f'results/{instance_num}/optimal_result.xlsx'
    else:
        filepath=f'results/{instance_num}/result.xlsx'

    with pd.ExcelWriter(filepath) as writer:
        U_unit.to_excel(writer, sheet_name='机组状态', index=False)
        P_unit.to_excel(writer, sheet_name='机组发电功率', index=False)
        ES_status.to_excel(writer, sheet_name='储能状态', index=False)
        ES_power.to_excel(writer, sheet_name='储能充放电功率', index=False)
    
    addColors(filepath,bin_table_name=['机组状态','储能状态'],continuous_table_name=['机组发电功率','储能充放电功率'])

def readSols(filePath):
    '''
    读取solution.sol文件，返回决策变量的值
    '''

    with open(filePath, 'r') as f:
        lines = f.readlines()
        
    lines=lines[1:]
    storage_status_lines = []
    storage_power_lines = []
    unit_status_lines = []
    unit_power_lines = []
    
    for line in lines:
        if line.startswith("storage") and "_s_" in line:
            storage_status_lines.append(line)
        elif line.startswith("storage") and "_p_" in line:
            storage_power_lines.append(line)
        elif line.startswith("unit") and "_s_" in line:
            unit_status_lines.append(line)
        elif line.startswith("unit") and "_p_" in line:
            unit_power_lines.append(line)
            
    #读取储能的信息
    N_ESs=int(len(storage_status_lines)/24)
    P_ES_ch=np.zeros((N_ESs,24))
    P_ES_dch=np.zeros((N_ESs,24))
    U_ES=np.zeros((N_ESs,24))
    U_ch=np.zeros((N_ESs,24))
    U_dch=np.zeros((N_ESs,24))

    
    #储能充放电功率
    for t,line in enumerate(storage_power_lines):
        line=line.split()
        P_ES=float(line[1])
        if float(line[1])<=0:
            P_ES_ch[t//24,t%24]=-P_ES
            P_ES_dch[t//24,t%24]=0
        else:
            P_ES_ch[t//24,t%24]=0
            P_ES_dch[t//24,t%24]=P_ES
            
    #储能状态
    for t,line in enumerate(storage_status_lines):
        line=line.split()

        if line[1]=='-1': #充电状态
            U_ES[t//24,t%24]=0
            U_ch[t//24,t%24]=1
            U_dch[t//24,t%24]=0
            
        elif line[1]=='1': #放电状态
            U_ES[t//24,t%24]=0
            U_ch[t//24,t%24]=0
            U_dch[t//24,t%24]=1
        
        elif line[1]=='0': #停机状态
            U_ES[t//24,t%24]=1
            U_ch[t//24,t%24]=0
            U_dch[t//24,t%24]=0
        
        else:
            raise ValueError("Invalid storage status")
            
    #读取机组的信息
    N_unit=int(len(unit_power_lines)/24)
    P_unit=np.zeros((N_unit,24))
    U_unit=np.zeros((N_unit,24))
    
    #机组发电功率
    for t,line in enumerate(unit_power_lines):
        line=line.split()
        P_unit[t//24,t%24]=float(line[1])
        
    #机组状态
    for t,line in enumerate(unit_status_lines):
        line=line.split()
        if line[1]=='1': #发电状态
            U_unit[t//24,t%24]=1
        elif line[1]=='0':
            U_unit[t//24,t%24]=0
        else:
            raise ValueError("Invalid unit status")
        
    Vars={"P_ES_ch":P_ES_ch,
              "P_ES_dch":P_ES_dch,
              "U_ES":U_ES,
              "U_ch":U_ch,
              "U_dch":U_dch,
              "P_unit":P_unit,
              "U_unit":U_unit}
    
    return Vars


def writeSols(model,U_unit,P_unit,U_ch,U_dch,U_ES,P_ES_ch,P_ES_dch,instance_num):
    '''
    将求解结果写入solution.sol文件
    输入：
    model: 求解模型
    U_unit: 机组状态
    P_unit: 机组发电功率
    U_ch: 储能充电状态
    U_dch: 储能放电状态
    U_ES: 储能状态
    P_ES_ch: 储能充电功率
    P_ES_dch: 储能放电功率
    instance_num: 实例编号
    '''
    
    #检测路径，如果不存在则创建
    if not os.path.exists(f'results/{instance_num}'):
        os.makedirs(f'results/{instance_num}')
        
    N_ESs=U_ch.shape[0]
    N_units=U_unit.shape[0]
    eps=1e-2
    
    # 保存结果
    with open(f'results/{instance_num}/solution.sol', 'w') as f:
        f.write(f"# Objective value = {model.objval}\n")
        
        #写入储能的结果
        for i in range(N_ESs):
            for t in range(24):
                if abs(U_ch[i,t]-1)<eps:
                    f.write(f"storage{i+1}_s_{t} -1\n")
                    f.write(f"storage{i+1}_p_{t} {-P_ES_ch[i,t]}\n")
                elif abs(U_dch[i,t]-1)<eps:
                    f.write(f"storage{i+1}_s_{t} 1\n")
                    f.write(f"storage{i+1}_p_{t} {P_ES_dch[i,t]}\n")
                elif abs(U_ES[i,t]-1)<eps:
                    f.write(f"storage{i+1}_s_{t} 0\n")
                    f.write(f"storage{i+1}_p_{t} 0\n")
                else:
                    raise ValueError("Invalid storage status") #确保每个决策变量都write进去了
                    
        #写入机组的结果
        for i in range(N_units):
            for t in range(24):
                if abs(U_unit[i,t]-1)<eps:
                    f.write(f"unit{i+1}_s_{t} 1\n")
                    f.write(f"unit{i+1}_p_{t} {P_unit[i,t]}\n")
                elif abs(U_unit[i,t]-0)<eps:
                    f.write(f"unit{i+1}_s_{t} 0\n")
                    f.write(f"unit{i+1}_p_{t} 0\n")

                else:
                    raise ValueError("Invalid unit status")
                    
    print(f"Results have been written to results/{instance_num}/solution.sol")

def checkConstraints(model):
    
    constraints = model.getConstrs()
    pass



def selfCheckLoadBalance(instance_num,eps=1e-4):

    filepath=f"results/{instance_num}/solution.sol"
    Vars=readSols(filepath)
    load=txt_to_dataframe(read_txt(f'data/instances/{instance_num}/slf.txt'))
    for t in range(24):
        P_load=np.sum(Vars["P_unit"][:,t])+np.sum(Vars["P_ES_dch"][:,t])-np.sum(Vars["P_ES_ch"][:,t])
        if abs(P_load-load['系统负荷大小（MW）'][t])<eps:
            print (f"负荷平衡约束在时刻{t}满足")

        else:
            print (f"负荷平衡约束在时刻{t}不满足, 当前负荷为{P_load}, 系统负荷为{load['系统负荷大小（MW）'][t]}")
        
#selfCheckLoadBalance(60)
#selfCheckLoadBalance(200)
