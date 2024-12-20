import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from io import StringIO


def read_txt(file_path):

    with open(file_path, 'r', encoding='ANSI') as file:
        lines = file.readlines()

    #去掉开头的"//"和空格的影响
    lines = [line.strip().lstrip('//') for line in lines]

    #修复unitdata.txt文件中的表头格式问题
    if 'unitdata.txt' in file_path:
        if '下爬坡率 (MW/h)' in lines[0]:
            lines[0] = lines[0].replace('下爬坡率 (MW/h)', '下爬坡率(MW/h)')

    return lines


def txt_to_dataframe(lines):

    cleaned_lines = []
    for line in lines:
        cleaned_line = '\t'.join(line.split())
        cleaned_lines.append(cleaned_line)

    cleaned_data = "\n".join(cleaned_lines)

    df = pd.read_csv(StringIO(cleaned_data), delimiter='\t')

    return df


def getSols(filePath):

    with open(filePath, 'r') as f:
        lines = f.readlines()
        
    lines=lines[1:]
    storage_status_lines = []
    storage_power_lines = []
    unit_status_lines = []
    unit_power_lines = []
    
    for line in lines:
        if line.startswith("storage") and "_s_" in line:
            storage_status_lines.append(line)
        elif line.startswith("storage") and "_p_" in line:
            storage_power_lines.append(line)
        elif line.startswith("unit") and "_s_" in line:
            unit_status_lines.append(line)
        elif line.startswith("unit") and "_p_" in line:
            unit_power_lines.append(line)
            
    #读取储能的信息
    P_ES_ch=np.zeros(24)
    P_ES_dch=np.zeros(24)
    U_ES=np.zeros(24)
    U_ch=np.zeros(24)
    U_dch=np.zeros(24)
    
    #储能充放电功率
    for t,line in enumerate(storage_power_lines):
        line=line.split()
        if float(line[1])<=0:
            P_ES_ch[t]=float(line[1])
            P_ES_dch[t]=0
        else:
            P_ES_ch[t]=0
            P_ES_dch[t]=float(line[1])
            
    #储能状态
    for t,line in enumerate(storage_status_lines):
        line=line.split()
        if line[1]=='-1': #充电状态
            U_ES[t]=0
            U_ch[t]=1
            U_dch[t]=0
            
        elif line[1]=='1': #放电状态
            U_ES[t]=0
            U_ch[t]=0
            U_dch[t]=1
        
        elif line[1]=='0': #停机状态
            U_ES[t]=1
            U_ch[t]=0
            U_dch[t]=0
            
    #读取机组的信息
    N_unit=int(len(unit_power_lines)/24)
    P_unit=np.zeros((N_unit,24))
    U_unit=np.zeros((N_unit,24))
    
    #机组发电功率
    for t,line in enumerate(unit_power_lines):
        #t/24取余数为机组编号
        line=line.split()
        P_unit[t//24,t%24]=float(line[1])
        
    #机组状态
    for t,line in enumerate(unit_status_lines):
        line=line.split()
        if line[1]=='1': #发电状态
            U_unit[t//24,t%24]=1
        elif line[1]=='0':
            U_unit[t//24,t%24]=0
        
        
        
    Vars={"P_ES_ch":P_ES_ch,
              "P_ES_dch":P_ES_dch,
              "U_ES":U_ES,
              "U_ch":U_ch,
              "U_dch":U_dch,
              "P_unit":P_unit,
              "U_unit":U_unit}
    
    return Vars
            
            
filePath = "data/instances/1/solution.sol"   
Vars=getSols(filePath)
load= txt_to_dataframe(read_txt('data/instances/1/slf.txt'))


U_unit=pd.DataFrame(Vars["U_unit"])
U_unit['机组编号']=U_unit.index+1
cols = list(U_unit)
cols.insert(0, cols.pop(cols.index('机组编号')))
U_unit = U_unit.loc[:, cols]

P_unit=pd.DataFrame(Vars["P_unit"])
P_unit['机组编号']=P_unit.index+1
cols = list(P_unit)
cols.insert(0, cols.pop(cols.index('机组编号')))
P_unit = P_unit.loc[:, cols]
P_unit["负荷"]=load['系统负荷大小（MW）']

ES_status = pd.DataFrame()
ES_status['充电状态'] = Vars["U_ch"]
ES_status['放电状态'] = Vars["U_dch"]
ES_status['停机状态'] = Vars["U_ES"]
ES_status['充电功率'] = Vars["P_ES_ch"]
ES_status['放电功率'] = Vars["P_ES_dch"]

with pd.ExcelWriter('results/optimal_result.xlsx') as writer:
    U_unit.to_excel(writer, sheet_name='机组状态', index=False)
    P_unit.to_excel(writer, sheet_name='机组发电功率', index=False)
    ES_status.to_excel(writer, sheet_name='储能状态', index=False)

# 加载工作簿
wb = load_workbook('results/optimal_result.xlsx')
ws = wb['机组状态']

# 定义填充样式
fill_1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
fill_0 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色

# 更改单元格底色
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=ws.max_column):
    for cell in row:
        if cell.value == 1:
            cell.fill = fill_1
        elif cell.value == 0:
            cell.fill = fill_0

# 为连续值的变量设置颜色渐变
ws = wb['机组发电功率']
color_scale_rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                  mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                  end_type='max', end_color='FF0000')
ws.conditional_formatting.add(f'B2:Z{ws.max_row}', color_scale_rule)


# 保存工作簿
wb.save('results/optimal_result.xlsx')

