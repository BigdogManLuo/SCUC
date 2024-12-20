import pandas as pd
from io import StringIO
import re
import cvxpy as cp
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule


def read_txt(file_path):

    with open(file_path, 'r', encoding='ANSI') as file:
        lines = file.readlines()

    #去掉开头的"//"和空格的影响
    lines = [line.strip().lstrip('//') for line in lines]

    #修复unitdata.txt文件中的表头格式问题
    if 'unitdata.txt' in file_path:
        if '下爬坡率 (MW/h)' in lines[0]:
            lines[0] = lines[0].replace('下爬坡率 (MW/h)', '下爬坡率(MW/h)')

    return lines


def txt_to_dataframe(lines):

    cleaned_lines = []
    for line in lines:
        cleaned_line = '\t'.join(line.split())
        cleaned_lines.append(cleaned_line)

    cleaned_data = "\n".join(cleaned_lines)

    df = pd.read_csv(StringIO(cleaned_data), delimiter='\t')

    return df


def readStorageBidInfo(file_path):

    lines=read_txt(file_path)
    
    cleaned_lines = lines[1:]
    cleaned_lines = "\t".join(cleaned_lines)
    data = list(map(int, cleaned_lines.split()))
    
    data = data[1:]

    return data


def parse_log_file(file_path):

    with open(file_path, 'r', encoding='ANSI') as file:
        content = file.read()

    # 使用正则表达式提取每个数据块
    blocks = {
        'BranchUnitSensi': re.search(r'<BranchUnitSensi::dky type=全数>(.*?)</BranchUnitSensi::dky>', content, re.DOTALL),
        'BranchData': re.search(r'<BranchData::dky type=全数>(.*?)</BranchData::dky>', content, re.DOTALL),
        'BranchBase': re.search(r'<BranchBase::dky type=全数>(.*?)</BranchBase::dky>', content, re.DOTALL)
    }

    df_tuple = ()

    # 提取每个数据块并转换为DataFrame
    for block_name, block in blocks.items():
        if block:
            block_content = block.group(1).strip()
            lines = block_content.splitlines()
            df=txt_to_dataframe(lines)
            df = df.loc[:, ~df.columns.str.contains('//')]
            df_tuple += (df,)

    return df_tuple

def getSegmentedCostInfo(prices,deltaP,P_min):
   
    # 计算每个区间的斜率和截距
    a = []
    b = []
    for i in range(len(prices) - 1):
        # 计算每个区间的斜率
        slope = (prices[i+1] - prices[i]) / deltaP  # 斜率
        intercept = prices[i] - slope * (P_min + i * deltaP)  # 截距
        a.append(slope)
        b.append(intercept)

    return a,b

def getSegmentedPoints(num_seg,deltaP,P_min):

    # 计算每段区间的左右端点

    Pmins=[]
    Pmaxs=[]
    
    for i in range(num_seg):
        Pmins.append(P_min+i*deltaP)
        Pmaxs.append(P_min+(i+1)*deltaP)


    return Pmins,Pmaxs


def getSols(filePath):

    with open(filePath, 'r') as f:
        lines = f.readlines()
        
    lines=lines[1:]
    storage_status_lines = []
    storage_power_lines = []
    unit_status_lines = []
    unit_power_lines = []
    
    for line in lines:
        if line.startswith("storage") and "_s_" in line:
            storage_status_lines.append(line)
        elif line.startswith("storage") and "_p_" in line:
            storage_power_lines.append(line)
        elif line.startswith("unit") and "_s_" in line:
            unit_status_lines.append(line)
        elif line.startswith("unit") and "_p_" in line:
            unit_power_lines.append(line)
            
    #读取储能的信息
    P_ES_ch=np.zeros(24)
    P_ES_dch=np.zeros(24)
    U_ES=np.zeros(24)
    U_ch=np.zeros(24)
    U_dch=np.zeros(24)
    
    #储能充放电功率
    for t,line in enumerate(storage_power_lines):
        line=line.split()
        if float(line[1])<=0:
            P_ES_ch[t]=-float(line[1])
            P_ES_dch[t]=0
        else:
            P_ES_ch[t]=0
            P_ES_dch[t]=float(line[1])
            
    #储能状态
    for t,line in enumerate(storage_status_lines):
        line=line.split()
        if line[1]=='-1': #充电状态
            U_ES[t]=0
            U_ch[t]=1
            U_dch[t]=0
            
        elif line[1]=='1': #放电状态
            U_ES[t]=0
            U_ch[t]=0
            U_dch[t]=1
        
        elif line[1]=='0': #停机状态
            U_ES[t]=1
            U_ch[t]=0
            U_dch[t]=0
            
    #读取机组的信息
    N_unit=int(len(unit_power_lines)/24)
    P_unit=np.zeros((N_unit,24))
    U_unit=np.zeros((N_unit,24))
    
    #机组发电功率
    for t,line in enumerate(unit_power_lines):
        #t/24取余数为机组编号
        line=line.split()
        P_unit[t//24,t%24]=float(line[1])
        
    #机组状态
    for t,line in enumerate(unit_status_lines):
        line=line.split()
        if line[1]=='1': #发电状态
            U_unit[t//24,t%24]=1
        elif line[1]=='0':
            U_unit[t//24,t%24]=0
        
    Vars={"P_ES_ch":P_ES_ch,
              "P_ES_dch":P_ES_dch,
              "U_ES":U_ES,
              "U_ch":U_ch,
              "U_dch":U_dch,
              "P_unit":P_unit,
              "U_unit":U_unit}
    
    return Vars

def calculateCosts(Vars):

    P_ES_ch=Vars['P_ES_ch']
    P_ES_dch=Vars['P_ES_dch']
    U_ES=Vars['U_ES']
    U_ch=Vars['U_ch']
    U_dch=Vars['U_dch']
    P_unit=Vars['P_unit']
    U_unit=Vars['U_unit']

    #充电成本
    Cost_ES_ch=-stbidprice[0]*np.sum(P_ES_ch)

    #放电成本
    Cost_ES_dch=0
    for t in range(24):
        
        #判断P_ES_ch属于哪一段
        flag=0
        if P_ES_dch[t]!=0:
            for j in range(len(stbidcapactiy)-1):
                if P_ES_dch[t]>=P_ES_mins[j] and P_ES_dch[t]<=P_ES_maxs[j]:
                    flag=1
                    break
            if flag==0:
                raise ValueError('P_ES_dch[t]不在任何一段范围内')

            Cost_ES_dch+=a_ES[j]*P_ES_dch[t]+b_ES[j]*U_dch[t]
            
        else:
            Cost_ES_dch+=0
        

    #通过Unit推导启动和停机的indicator
    V_unit=np.zeros_like(U_unit)
    W_unit=np.zeros_like(U_unit)

    for i in range(U_unit.shape[0]):
        for t in range(24):
            if t==0: #初始时刻的启停机状态
                V_unit[i,t]=U_unit[i,t]-unitdata['初始状态(1开机,0停机)'][i]
                W_unit[i,t]=unitdata['初始状态(1开机,0停机)'][i]-U_unit[i,t]

            if t>0:
                V_unit[i,t]=max(0,U_unit[i,t]-U_unit[i,t-1])
                W_unit[i,t]=max(0,U_unit[i,t-1]-U_unit[i,t])

    #机组启动成本
    Cost_unit_start=0
    for t in range(24):
        Cost_unit_start+=np.sum(unitdata['启动成本（元）'].values*U_unit[:,t])

    #机组运行成本
    Cost_unit_opr=0
    for t in range(24):
        for i in range(P_unit.shape[0]):

            #计算成本区间上下限
            P_unit_mins,P_unit_maxs=getSegmentedPoints(num_seg=bid_capacity.shape[1]-1,deltaP=bid_capacity.iloc[i,1],P_min=unitdata['最小出力(MW)'][i])

            #每个区间的斜率和截距
            a_unit,b_unit=getSegmentedCostInfo(prices=bid_price.iloc[i,2:],deltaP=bid_capacity.iloc[i,1],P_min=unitdata['最小出力(MW)'][i])

            #判断P_unit[i,t]属于哪一段
            flag=0
            if P_unit[i,t]!=0:
                for j in range(bid_capacity.shape[1]-1):
                    if P_unit[i,t]>=P_unit_mins[j] and P_unit[i,t]<=P_unit_maxs[j]:
                        flag=1
                        break
                    
                if flag==0:
                    raise ValueError('P_unit[i,t]不在任何一段范围内')
                
                Cost_unit_opr+=a_unit[j]*P_unit[i,t]+b_unit[j]*U_unit[i,t]
            else:
                Cost_unit_opr+=0
            

    return Cost_ES_ch,Cost_ES_dch,Cost_unit_start,Cost_unit_opr

def addColors(filePath,bin_table_name=['机组状态','机组启动状态','机组停机状态','储能状态'],continuous_table_name=['机组功率']):
    # 加载工作簿
    wb = load_workbook(filePath)

    # 定义填充样式
    fill_1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
    fill_0 = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色


    for sheet_name in bin_table_name:

        ws = wb[sheet_name]

        # 更改单元格底色
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=ws.max_column):
            for cell in row:
                if cell.value == 1:
                    cell.fill = fill_1
                elif cell.value == 0:
                    cell.fill = fill_0


    # 为连续值的变量设置颜色渐变
    for sheet_name in continuous_table_name:

        ws = wb[sheet_name]

        color_scale_rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                        mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                        end_type='max', end_color='FF0000')
        ws.conditional_formatting.add(f'B2:Z{ws.max_row}', color_scale_rule)


    # 保存工作簿
    wb.save(filePath)



#---------------------------Parameters-----------------------------------------#
bid_capacity = txt_to_dataframe(read_txt('data/instances/1/bidcapacity.txt'))
bid_price= txt_to_dataframe(read_txt('data/instances/1/bidprice.txt'))
section= txt_to_dataframe(read_txt('data/instances/1/section.txt'))
load= txt_to_dataframe(read_txt('data/instances/1/slf.txt'))
stbidcapactiy= readStorageBidInfo('data/instances/1/stbidcapacity.txt')
stbidprice= readStorageBidInfo('data/instances/1/stbidprice.txt')
storagebasic= txt_to_dataframe(read_txt('data/instances/1/storagebasic.txt'))
unitdata= txt_to_dataframe(read_txt('data/instances/1/unitdata.txt'))
gen_senses,load_sense,branch=parse_log_file('data/instances/1/branch_1.log')


M=100000 #大M法中的大M值
T=24 

#----------------------储能分段线性信息---------------------------
deltaP_ES=stbidcapactiy[1] #储能分段功率的间隔
a_ES,b_ES=getSegmentedCostInfo(stbidprice[2:],deltaP=deltaP_ES,P_min=0) #储能成本分段线性化
P_ES_mins,P_ES_maxs=getSegmentedPoints(num_seg=len(stbidcapactiy)-2,deltaP=stbidcapactiy[1],P_min=storagebasic['最小发电功率（MW）'][0]) #储能分段功率的上下限


#---------------------断面信息-------------------------
restricted_branches = section['断面组成']
PF_load=np.zeros((len(restricted_branches),T)) #负荷对断面的潮流

#计算每个时刻每个断面的负荷潮流
for t in range(T):
    for j in range(len(restricted_branches)):
        
        #找到对应的断面在load_sense中的部分
        load_in_branch=load_sense[load_sense['支路名称（ID）']==restricted_branches[j]].reset_index(drop=True)
        
        #负荷对断面的潮流
        PF_load[j,t]=load_in_branch['母线负荷对该支路潮流的灵敏度值乘积和'][t]

#%%
#---------------------------Results 校验-----------------------------#
Vars=getSols('data/instances/1/solution.sol')
Cost_ES_ch,Cost_ES_dch,Cost_unit_start,Cost_unit_opr=calculateCosts(Vars)


#%%
#---------------------------Variables-----------------------------#

#Unit Data
P_unit=cp.Variable((unitdata.shape[0],T))
U_unit=cp.Variable((unitdata.shape[0],T),boolean=True) #开停机state
V_unit=cp.Variable((unitdata.shape[0],T),boolean=True) #启动indicator
W_unit=cp.Variable((unitdata.shape[0],T),boolean=True) #停机indicator
PF_unit=cp.Variable((len(restricted_branches),T)) #火电机组对断面的潮流
L_unit=[cp.Variable((unitdata.shape[0],T),boolean=True) for j in range(len(bid_capacity)-1)] #火电机组属于哪个分段indicator
Cost_unit_start=cp.Variable((unitdata.shape[0],T))
Cost_unit_down=cp.Variable((unitdata.shape[0],T)) #虚拟的停机成本用于松弛停机indicator
Cost_unit_opr=cp.Variable((unitdata.shape[0],T))
Cost_unit_opr_per_seg=[cp.Variable((unitdata.shape[0],T)) for j in range(len(bid_capacity)-1)] #火电机组成本分段线性化变量


#Storage Data
U_ch = cp.Variable((storagebasic.shape[0], T), boolean=True) #充电state
U_dch = cp.Variable((storagebasic.shape[0], T), boolean=True) #放电state
U_ES=cp.Variable((storagebasic.shape[0],T),boolean=True) #停机state
Z_ch=cp.Variable((storagebasic.shape[0],T),boolean=True) #充电indicator
Z_dch=cp.Variable((storagebasic.shape[0],T),boolean=True) #放电indicator
Z_ES=cp.Variable((storagebasic.shape[0],T),boolean=True) #停机indicator
P_ES_ch=cp.Variable((storagebasic.shape[0],T)) #充电功率
P_ES_dch=cp.Variable((storagebasic.shape[0],T)) #放电功率
PF_ES=cp.Variable((len(restricted_branches),T)) #储能机组对断面的潮流
ES=cp.Variable((storagebasic.shape[0],T+1)) #储能容量剩余
L_ES=[cp.Variable((storagebasic.shape[0],T),boolean=True) for j in range(len(stbidcapactiy)-2)] #储能属于哪个分段indicator
Cost_ES_ch=cp.Variable((storagebasic.shape[0],T)) #储能充电成本
Cost_ES_dch=cp.Variable((storagebasic.shape[0],T)) #储能放电成本
Cost_ES_dch_per_seg=[cp.Variable((storagebasic.shape[0],T)) for j in range(len(stbidcapactiy)-2)] #储能成本分段线性化变量

#---------------------------Constraints-----------------------------#
constraints=[]

#火电机组约束
for i in range(unitdata.shape[0]):

    #计算成本区间上下限
    P_unit_mins,P_unit_maxs=getSegmentedPoints(num_seg=bid_capacity.shape[1]-1,deltaP=bid_capacity.iloc[i,1],P_min=unitdata['最小出力(MW)'][i])

    #每个区间的斜率和截距
    a_unit,b_unit=getSegmentedCostInfo(prices=bid_price.iloc[i,2:],deltaP=bid_capacity.iloc[i,1],P_min=unitdata['最小出力(MW)'][i])

    for t in range(T):

        #发电功率上下限约束
        constraints+=[P_unit[i,t]>=unitdata['最小出力(MW)'][i]*U_unit[i,t]]
        constraints+=[P_unit[i,t]<=unitdata['最大出力(MW)'][i]*U_unit[i,t]]

        #最小开机时间约束
        if unitdata['初始状态(1开机,0停机)'][i]==1: #如果初始状态为开机
            if unitdata['初始状态持续时间(h)'][i]-unitdata['最小开机时间(h)'][i]+t<0: #初始状态持续时间小于最小开机时间,则需要继续保持初始状态
                constraints+=[U_unit[i,t]==1]
        
        if t+unitdata['最小开机时间(h)'][i]<=T:
            constraints+=[sum([U_unit[i,k] for k in range(t,t+unitdata['最小开机时间(h)'][i])])>=unitdata['最小开机时间(h)'][i]*V_unit[i,t]]
        else:
            constraints+=[sum([U_unit[i,k] for k in range(t,T)])>=V_unit[i,t]*(T-t)]

        
        #最小停机时间约束
        if unitdata['初始状态(1开机,0停机)'][i]==0:    #如果初始状态为停机
            if unitdata['初始状态持续时间(h)'][i]-unitdata['最小停机时间(h)'][i]+t<0: #初始状态持续时间小于最小停机时间,则需要继续保持初始状态
                constraints+=[U_unit[i,t]==0]
        
        if t+unitdata['最小停机时间(h)'][i]<=T:
            constraints+=[sum([1-U_unit[i,k] for k in range(t,t+unitdata['最小停机时间(h)'][i])])>=unitdata['最小停机时间(h)'][i]*W_unit[i,t]]
        else:
            constraints+=[sum([1-U_unit[i,k] for k in range(t,T)])>=W_unit[i,t]*(T-t)]
        
        #启停机状态转换
        if t==0: #初始时刻的启停机状态
            constraints+=[V_unit[i,t]>=U_unit[i,t]-unitdata['初始状态(1开机,0停机)'][i]]
            constraints+=[W_unit[i,t]>=unitdata['初始状态(1开机,0停机)'][i]-U_unit[i,t]]

        if t>0:
            constraints+=[V_unit[i,t]>=U_unit[i,t]-U_unit[i,t-1]]
            constraints+=[W_unit[i,t]>=U_unit[i,t-1]-U_unit[i,t]]
        
        #爬坡率约束 （开停机动作可以不满足爬坡率约束）
        if t==0: 
            constraints+=[P_unit[i,t]-unitdata['初始出力(MW)'][i]<=unitdata['上爬坡率(MW/h)'][i]+M*(V_unit[i,t])]
            constraints+=[unitdata['初始出力(MW)'][i]-P_unit[i,t]<=unitdata['下爬坡率(MW/h)'][i]+M*(W_unit[i,t])]
        
        if t>0:
            constraints+=[P_unit[i,t]-P_unit[i,t-1]<=unitdata['上爬坡率(MW/h)'][i]+M*(V_unit[i,t])]
            constraints+=[P_unit[i,t-1]-P_unit[i,t]<=unitdata['下爬坡率(MW/h)'][i]+M*(W_unit[i,t])]
        
        #启动成本
        constraints+=[Cost_unit_start[i,t]==unitdata['启动成本（元）'][i]*V_unit[i,t]]

        #虚拟的停机成本
        constraints+=[Cost_unit_down[i,t]==W_unit[i,t]]
        
        #分段功率上下限约束
        constraints+=[P_unit[i,t]>=sum(P_unit_mins[j]*L_unit[j][i,t] for j in range(bid_capacity.shape[1]-1))]
        constraints+=[P_unit[i,t]<=sum(P_unit_maxs[j]*L_unit[j][i,t] for j in range(bid_capacity.shape[1]-1))]

        #有且只有一段被激活,均不激活的时候为P_unit为0
        constraints+=[sum(L_unit[j][i,t] for j in range(len(bid_capacity)-1))<=1]

        #运行成本
        for j in range(bid_capacity.shape[1]-1):

            constraints+=[Cost_unit_opr_per_seg[j][i,t]>=0]
            constraints+=[Cost_unit_opr_per_seg[j][i,t]<=L_unit[j][i,t]*M]
            constraints+=[Cost_unit_opr_per_seg[j][i,t]<=a_unit[j]*P_unit[i,t]+b_unit[j]*U_unit[i,t]]
            constraints+=[Cost_unit_opr_per_seg[j][i,t]>=a_unit[j]*P_unit[i,t]+b_unit[j]*U_unit[i,t]-M*(1-L_unit[j][i,t])]

        #最终运行成本
        constraints+=[Cost_unit_opr[i,t]==sum(Cost_unit_opr_per_seg[j][i,t] for j in range(bid_capacity.shape[1]-1))]
        

#储能约束
for i in range(storagebasic.shape[0]):

    # 初始容量
    constraints += [ES[i, 0] == storagebasic['初始容量（MWh）'][i]]
    
    # 终止容量
    constraints += [ES[i, T] >= storagebasic['终止容量（MWh）'][i]]

    for t in range(T):

        #充电功率约束
        constraints += [P_ES_ch[i, t] == storagebasic['抽水固定功率（MW）'][i] * U_ch[i, t]]
        
        #放电功率约束
        constraints += [P_ES_dch[i, t] >= storagebasic['最小发电功率（MW）'][i]* U_dch[i, t]]
        constraints += [P_ES_dch[i, t] <=storagebasic['最大发电功率（MW）'][i]* U_dch[i, t]]
        
        #储能容量约束
        constraints += [ES[i, t+1] == ES[i, t] + P_ES_ch[i, t] - P_ES_dch[i, t]]
        constraints += [ES[i, t+1] <= storagebasic['最大容量（MWh）'][i]]
        constraints += [ES[i, t+1] >= 0]
        
        #最小抽水时间约束
        if t+storagebasic['最小抽水时段'][i]<=T:
            constraints += [sum([U_ch[i, k] for k in range(t, t + storagebasic['最小抽水时段'][i])]) >= storagebasic['最小抽水时段'][i] * Z_ch[i, t]]
        else:
            constraints += [sum([U_ch[i, k] for k in range(t, T)]) >= Z_ch[i, t] * (T - t)]

        #最小放电时间约束
        if t+storagebasic['最小发电时段'][i]<=T:
            constraints += [sum([U_dch[i, k] for k in range(t, t + storagebasic['最小发电时段'][i])]) >= storagebasic['最小发电时段'][i] * Z_dch[i, t]]
        else:
            constraints += [sum([U_dch[i, k] for k in range(t, T)]) >= Z_dch[i, t] * (T - t)]

        #最小停机时间约束
        if t+storagebasic['最小停机时段'][i]<=T:
            constraints += [sum([U_ES[i, k] for k in range(t, t + storagebasic['最小停机时段'][i])]) >= storagebasic['最小停机时段'][i] * Z_ES[i, t]]
        else:
            constraints += [sum([U_ES[i, k] for k in range(t, T)]) >= Z_ES[i, t] * (T - t)]

        #状态互斥约束
        constraints += [U_ch[i, t] + U_dch[i, t] + U_ES[i, t] == 1]

        #状态转换约束
        if t==0:
            constraints += [Z_dch[i, t] == U_dch[i, t]-0] #如果第一时刻选择放电，则放电indicator激活
            constraints += [Z_ch[i, t] == U_ch[i, t]-0]  #如果第一时刻选择充电，则充电indicator激活
            constraints += [Z_ES[i, t] == U_ES[i, t]-1] #第一时刻停机也不会激活停机indicator  

        if t > 0:
            constraints += [Z_ch[i, t] >= U_ch[i, t] - U_ch[i, t - 1]]
            constraints += [Z_dch[i, t] >= U_dch[i, t] - U_dch[i, t - 1]]
            constraints += [Z_ES[i, t] >= U_ES[i, t] - U_ES[i, t - 1]]


        #状态切换，充电状态与发电状态相互的切换必须经过停机状态
        if t > 0:
            constraints += [Z_dch[i, t] <= U_ES[i, t - 1]]
            constraints += [Z_ch[i, t] <= U_ES[i, t - 1]]

        #充电成本
        constraints+= [Cost_ES_ch[i,t]==-stbidprice[0]]  
        
        
        #放电功率分段上下限约束
        constraints+=[P_ES_dch[i,t]>=sum(P_ES_mins[j]*L_ES[j][i,t] for j in range(len(stbidcapactiy)-2))]
        constraints+=[P_ES_dch[i,t]<=sum(P_ES_maxs[j]*L_ES[j][i,t] for j in range(len(stbidcapactiy)-2))]


        #有且只有一段被激活
        constraints+=[sum(L_ES[j][i,t] for j in range(len(stbidcapactiy)-2))<=1]
        
        
        #放电成本分段线性
        for j in range(len(stbidcapactiy)-2):
            
            constraints+=[Cost_ES_dch_per_seg[j][i,t]>=0]
            constraints+=[Cost_ES_dch_per_seg[j][i,t]<=L_ES[j][i,t]*M]
            constraints+=[Cost_ES_dch_per_seg[j][i,t]<=a_ES[j]*P_ES_dch[i,t]+b_ES[j]*U_dch[i,t]]
            constraints+=[Cost_ES_dch_per_seg[j][i,t]>=a_ES[j]*P_ES_dch[i,t]+b_ES[j]*U_dch[i,t]-M*(1-L_ES[j][i,t])]

        #最终放电成本
        constraints+=[Cost_ES_dch[i,t]==sum(Cost_ES_dch_per_seg[j][i,t] for j in range(len(stbidcapactiy)-2))]
        
        
#系统约束
for t in range(T):

    #系统平衡约束
    constraints+= [cp.sum(P_unit[:, t])+cp.sum(P_ES_dch[:, t])-cp.sum(P_ES_ch[:, t])==load['系统负荷大小（MW）'][t]]

    #系统备用约束
    constraints += [ cp.sum(unitdata['最大出力(MW)'].values-P_unit[:, t]) >= 0.1*load['系统负荷大小（MW）'][t]]
    
    #断面约束
    for j in range(len(restricted_branches)):

        #找到对应的断面在gen_senses中的位置
        branch_idx=gen_senses[gen_senses['支路中文名']==restricted_branches[j]].index[0]
        
        
        #火电机组对断面的潮流
        constraints+=[
            PF_unit[j,t]==sum(P_unit[i,t]*gen_senses[f'{i+1}对所列支路潮流的灵敏度值'][branch_idx] for i in range(unitdata.shape[0]))
        ]

        
        #储能机组对断面的潮流
        constraints+=[
            PF_ES[j,t]==(P_ES_dch[i,t]-P_ES_ch[i,t])*gen_senses[f'{len(unitdata)}对所列支路潮流的灵敏度值'][branch_idx]
        ]

        #断面潮流约束
        constraints+=[
            PF_unit[j,t]+PF_ES[j,t]-PF_load[j,t]<=section['断面限额'][j]
        ]
        
#---------------------------------Objective---------------------------------#
#目标函数=火电机组的启动成本+火电机组的运行成本+储能机组的运行成本
obj=cp.Minimize(cp.sum(Cost_unit_start)+cp.sum(Cost_unit_down)+
                cp.sum(Cost_unit_opr)+
                cp.sum(Cost_ES_ch)+cp.sum(Cost_ES_dch))


#---------------------------------Solve---------------------------------#
prob=cp.Problem(obj,constraints)
print('Solving...')
solver_opt={
    'solver':cp.GUROBI,
    'verbose':True,
    'MIPGap':0.0003}

prob.solve(**solver_opt)

if prob.status=='optimal':
    #---------------------------------Output---------------------------------#
    print('Optimal value:', prob.value)

    #机组状态
    Unit_status=pd.DataFrame()
    Unit_status['机组序号']=unitdata['机组序号']
    Unit_status['最小开机时间(h)']=unitdata['最小开机时间(h)']
    Unit_status['最小停机时间(h)']=unitdata['最小停机时间(h)']
    for t in range(T):
        Unit_status['第'+str(t+1)+'时刻状态']=U_unit.value[:,t]


    for i in range(unitdata.shape[0]):

        a_unit,b_unit=getSegmentedCostInfo(prices=bid_price.iloc[i,2:],deltaP=bid_capacity.iloc[i,1],P_min=unitdata['最小出力(MW)'][i])
        Unit_status.loc[i,'平均单位发电成本']=np.mean(a_unit[:])

    #机组启停机状态
    V_unit_status=pd.DataFrame()
    V_unit_status['机组序号']=unitdata['机组序号']
    for t in range(T):
        V_unit_status['第'+str(t+1)+'时刻启动']=V_unit.value[:,t]

    W_unit_status=pd.DataFrame()
    W_unit_status['机组序号']=unitdata['机组序号']
    for t in range(T):
        W_unit_status['第'+str(t+1)+'时刻停机']=W_unit.value[:,t]

    #机组功率
    P_unit_status=pd.DataFrame()
    P_unit_status['机组序号']=unitdata['机组序号']
    for t in range(T):
        P_unit_status['第'+str(t+1)+'时刻功率(MW)']=P_unit.value[:,t]
    P_unit_status['负荷']=load['系统负荷大小（MW）']

    #储能状态
    ES_status = pd.DataFrame()
    ES_status['储能容量(MWh)'] = ES.value[:,].flatten()
    ES_status['充电状态'] = np.append(U_ch.value.flatten(), 0)
    ES_status['放电状态'] = np.append(U_dch.value.flatten(), 0)
    ES_status['停机状态'] = np.append(U_ES.value.flatten(), 0)
    ES_status['充电功率(MW)'] = np.append(P_ES_ch.value.flatten(), 0)
    ES_status['放电功率(MW)'] = np.append(P_ES_dch.value.flatten(), 0)


    #写入到excel
    with pd.ExcelWriter('results/result.xlsx') as writer:
        Unit_status.to_excel(writer, sheet_name='机组状态', index=False)
        V_unit_status.to_excel(writer, sheet_name='机组启动状态', index=False)
        W_unit_status.to_excel(writer, sheet_name='机组停机状态', index=False)
        P_unit_status.to_excel(writer, sheet_name='机组功率', index=False)
        ES_status.to_excel(writer, sheet_name='储能状态', index=False)

    addColors('results/result.xlsx')

else:
    print('No optimal solution found')


''' 备注：
stbidcapactiy 是列表类型，用[x]来访问，其中x为索引
其余为pandas.DataFrame类型。

1. 索引DataFrame的某一列：
如果要访问其中某一列，可以用df['column_name']来访问，其中df为DataFrame对象，column_name为列名。例如：
bid_capacity['第三段功率(Mw)']

2. 索引DataFrame的某一行：
如果要访问其中某一行，可以用df.iloc[x]来访问，其中df为DataFrame对象，x为行索引。例如：
bid_capacity.iloc[2]

3. 索引DataFrame的某一列中的第x个元素：
用df['column_name'][x]来访问，其中df为DataFrame对象，column_name为列名，x为行索引。例如：
bid_capacity['第三段功率(Mw)'][2]，其中2就代表第3个机组

特别注意！！！！：python的索引是从0开始，因此在DataFrame中机组序号应该是索引值+1

'''


